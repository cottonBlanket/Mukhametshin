import cProfile
import concurrent
import csv
import os
import concurrent.futures as pool
import sys

import pandas as pd
import math
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, Side, Border
from openpyxl.utils import get_column_letter
import matplotlib.pyplot as plt
from jinja2 import Environment, FileSystemLoader
import pdfkit

currency_to_rub = {
    "AZN": 35.68,
    "BYR": 23.91,
    "EUR": 59.90,
    "GEL": 21.74,
    "KGS": 0.76,
    "KZT": 0.13,
    "RUR": 1,
    "UAH": 1.64,
    "USD": 60.66,
    "UZS": 0.0055,
}

def GetSalaryToRub(row):
    if math.isnan(row['salary_mean']):
        return 0
    return row['salary_mean'] * currency_to_rub[row['salary_currency']]

class DataSet():
    def __init__(self):
        self.folder_name = 'csv'
        self.inputValues = InputConect()
        self.vacancy = self.inputValues.professionName
        self.df = pd.read_csv(self.inputValues.fileName)


        self.df['salary_mean'] = self.df[['salary_from', 'salary_to']].mean(axis=1)
        # переводим зп в рубли
        self.df['salary'] = self.df.apply(GetSalaryToRub, axis=1)
        self.df['published_at'] = self.df['published_at'].apply(lambda x: int(x[:4]))
        df_vacancy = self.df[self.df['name'].str.contains(self.vacancy)]
        self.years = self.df['published_at'].unique()
        self.salaryByYears = {year: [] for year in self.years}
        self.vacsByYears = {year: 0 for year in self.years}
        self.vacSalaryByYears = {year: [] for year in self.years}
        self.vacCountByYears = {year: 0 for year in self.years}
        self.GetStaticByCities()
        self.initializeYearStatistics()

    def initializeYearStatistics(self):
        """Добавляет в словари статистик значения из файла
            Также запускает concurrent.futures
        """
        with concurrent.futures.ProcessPoolExecutor(max_workers=3) as executor:
            wait_complete = []
            for task in self.years:
                future = executor.submit(self.getStatisticByYear, task)
                wait_complete.append(future)

        for res in concurrent.futures.as_completed(wait_complete):
            result = res.result()
            year = result[0]
            self.salaryByYears[year] = result[1]
            self.vacsByYears[year] = result[2]
            self.vacSalaryByYears[year] = result[3]
            self.vacCountByYears[year] = result[4]

        self.salaryByYears = dict(sorted(self.salaryByYears.items()))
        self.vacsByYears = dict(sorted(self.vacsByYears.items()))
        self.vacSalaryByYears = dict(sorted(self.vacSalaryByYears.items()))
        self.vacCountByYears = dict(sorted(self.vacCountByYears.items()))
        self.PrintInfo()


    def getStatisticByYear(self, year):
        """Возвращает статистку за год в порядке:
            Среднее значение зарплаты за год,
            Количество вакансий за год.
            Среднее значение зарплаты за год для выбранной профессии.
            Количество вакансий за год для выбранной профессии
        """
        file_path = rf"{self.folder_name}\{year}.csv"
        if os.path.exists(file_path):
            df = pd.read_csv(file_path)
            df["salary_mean"] = df[["salary_from", "salary_to"]].mean(axis=1)
            df['salary'] = df.apply(GetSalaryToRub, axis=1)
            df1 = df[df["name"].str.contains(self.vacancy)]
            df_vacancy = df1[df1['area_name'] == self.inputValues.areaName]

            averageSalary = math.floor(df["salary"].mean())
            numberOfVacancies = len(df.index)
            averageSalaryProfession = 0 if df_vacancy.empty or math.isnan(df_vacancy['salary'].mean()) else math.floor(df_vacancy["salary"].mean())
            numberOfVacanciesProfession = 0 if df_vacancy.empty else len(df_vacancy.index)

            return [year, averageSalary, numberOfVacancies, averageSalaryProfession,
                                 numberOfVacanciesProfession]

    def PrintInfo(self):
        '''
        Вывод данных
        :return: void
        '''
        print('Динамика уровня зарплат по годам:', self.salaryByYears)
        print('Динамика количества вакансий по годам:', self.vacsByYears)
        print('Динамика уровня зарплат по годам для выбранной профессии:', self.vacSalaryByYears)
        print('Динамика количества вакансий по годам для выбранной профессии:', self.vacCountByYears)
        print('Уровень зарплат по городам (в порядке убывания):', self.salaryArea)
        print('Доля вакансий по городам (в порядке убывания):', self.countArea)
        Report(inputValues=self.inputValues, yearSalary=self.salaryByYears, yearSalary_needed=self.vacSalaryByYears,
               year_to_count=self.vacSalaryByYears, yearCount=self.vacCountByYears, areaSalary=self.salaryArea,
               areaPiece=self.countArea)

    def GetStaticByCities(self):
        '''
        Получение данных по городам, не требующих multiprocessing
        :return: void
        '''
        total = len(self.df)
        self.df['count'] = self.df.groupby('area_name')['area_name'].transform('count')
        # округление 1% по условию задачи
        df_norm = self.df[self.df['count'] > 0.01 * total]
        df_area = df_norm.groupby('area_name', as_index=False)['salary'].mean().sort_values(by='salary', ascending=False)
        df_count = df_norm.groupby('area_name', as_index=False)['count'].mean().sort_values(by='count', ascending=False)
        cities = df_count['area_name'].unique()

        self.salaryArea = {city: 0 for city in cities}
        self.countArea = {city: 0 for city in cities}
        for city in cities:
            self.salaryArea[city] = int(df_area[df_area['area_name'] == city]['salary'])
            self.salaryArea = dict(sorted(self.salaryArea.items(), key=lambda x: x[1], reverse=True))
            self.countArea[city] = round(int(df_count[df_count['area_name'] == city]['count']) / total, 4)

class Report:
    '''
    Класс для создания отчетов
    Attributes:
        inputValues (InputConect) входные данные
        yearSalary (dict) зп по годам
        yearSalary_needed (dict) зп по  годам для вакансии
        year_to_count (dict) ваканси по годам
        yearCount (dict) вакансии по годам для профессии
        areaSalary (dict) уровень зп по городам
        areaPiece (dict) доля вакансий по городам
    '''
    def __init__(self, inputValues, yearSalary, yearSalary_needed, year_to_count, yearCount, areaSalary, areaPiece):
        self.inputValues = inputValues
        self.yearSalary = yearSalary
        self.yearSalary_needed = yearSalary_needed
        self.year_to_count = year_to_count
        self.yearCount = yearCount
        self.areaSalary = areaSalary
        self.areaPiece = areaPiece
        self.generate_excel()
        self.generate_image()
        self.generate_pdf()

    def generate_excel(self):
        '''
        Создает файл отчета в экселе
        :return: void
        '''
        execelFile = Workbook()
        execelSheetFirst = execelFile.create_sheet(title="Статистика по годам", index=0)
        execelSheetFirst['A1'] = 'Год'
        execelSheetFirst['B1'] = 'Средняя зарплата'
        execelSheetFirst['C1'] = f'Средняя зарплата - {self.inputValues.professionName}'
        execelSheetFirst['D1'] = 'Количество вакансий'
        execelSheetFirst['E1'] = f'Количество вакансий - {self.inputValues.professionName}'
        execelSheetFirst['A1'].font = Font(bold=True)
        execelSheetFirst['B1'].font = Font(bold=True)
        execelSheetFirst['C1'].font = Font(bold=True)
        execelSheetFirst['D1'].font = Font(bold=True)
        execelSheetFirst['E1'].font = Font(bold=True)
        yearRow = list(self.yearSalary.keys())
        firstValue = list(self.yearSalary.values())
        secondValues = list(self.yearSalary_needed.values())
        thirdValues = list(self.year_to_count.values())
        fourthValues = list(self.yearCount.values())
        for i in range(0, 16):
            data = list()
            data.append(yearRow[i])
            data.append(firstValue[i])
            data.append(secondValues[i])
            data.append(thirdValues[i])
            data.append(fourthValues[i])
            execelSheetFirst.append(data)

        self.setBorder(columns=['A', 'B', 'C', 'D', 'E'], excelSheet=execelSheetFirst, numberSheet=0)
        self.setSizeColumns(excelSheet=execelSheetFirst)

        execelSheetSecond = execelFile.create_sheet(title="Статистика по городам", index=1)

        execelSheetSecond['A1'] = 'Город'
        execelSheetSecond['B1'] = 'Уровень зарплат'
        execelSheetSecond['D1'] = 'Город'
        execelSheetSecond['E1'] = 'Доля вакансий'
        execelSheetSecond['A1'].font = Font(bold=True)
        execelSheetSecond['B1'].font = Font(bold=True)
        execelSheetSecond['D1'].font = Font(bold=True)
        execelSheetSecond['E1'].font = Font(bold=True)
        cityes1 = list(self.areaSalary.keys())
        salaryes = list(self.areaSalary.values())
        cityes2 = list(self.areaPiece.keys())
        pieces = list(self.areaPiece.values())
        for i in range(0, 10):
            data = list()
            data.append(cityes1[i])
            data.append(salaryes[i])
            data.append("")
            data.append(cityes2[i])
            data.append(pieces[i])
            execelSheetSecond.append(data)

        self.setBorder(columns=['A', 'B', 'D', 'E'], excelSheet=execelSheetSecond, numberSheet=1)
        self.setSizeColumns(excelSheet=execelSheetSecond)

        execelFile.save('report.xlsx')

    def setSizeColumns(self, excelSheet):
        '''
        Устанавливает размер колонок
        :param excelSheet: инстанс листа эксель
        :return: void
        '''
        i = 0
        col_width = list()
        for col in excelSheet.columns:
            for j in range(len(col)):
                if j == 0:
                    col_width.append(len(str(col[j].value)))
                else:
                    if col_width[i] < len(str(col[j].value)):
                        col_width[i] = len(str(col[j].value))
            i = i + 1

        for i in range(len(col_width)):
            col_letter = get_column_letter(i + 1)
            if col_width[i] > 100:
                excelSheet.column_dimensions[col_letter].width = 100
            else:
                excelSheet.column_dimensions[col_letter].width = col_width[i] + 2

    def setBorder(self, columns, excelSheet, numberSheet):
        '''
        Устанавливает рамки в экселе
        :param columns: колонка в экселе
        :param excelSheet: инстанс листа в экселе
        :param numberSheet: номер листа
        :return: void
        '''
        side = Side(border_style='thin', color="00000000")
        for i in columns:
            column = excelSheet[i]
            for j in column:
                j.border = Border(left=side, right=side, top=side, bottom=side)
                if i == 'E' and numberSheet == 1:
                    j.number_format = openpyxl.styles.numbers.BUILTIN_FORMATS[10]

    def generate_image(self):
        '''
        Создает графики для аналитики c помощью matplotlib
        :return: void
        '''
        years = [i for i in self.yearSalary.keys()]
        yearSalary = [i for i in self.yearSalary_needed.values()]
        avg = [i for i in self.yearSalary.values()]
        fig = plt.figure()
        ax = fig.add_subplot(221)
        plt.rcParams.update({'font.size': 8})
        ax.set_title("Уровень зарплат по годам", fontsize=8)
        x_nums = range(len(self.yearSalary.keys()))
        width = 0.4
        x_list1 = list(map(lambda x: x - width / 2, x_nums))
        x_list2 = list(map(lambda x: x + width / 2, x_nums))
        ax.bar(x_list1, yearSalary, width, label="средняя з/п")
        ax.bar(x_list2, avg, width, label="з/п аналитик")
        ax.set_xticks(x_nums, years, rotation="vertical")
        ax.tick_params(axis="both", labelsize=8)
        ax.legend(fontsize=8)
        ax.grid(True, axis="y")

        ax = fig.add_subplot(222)
        ax.set_title("Количество вакансий по годам", fontsize=8)
        vacancies = [i for i in self.yearCount.values()]
        vacanciesAvg = [i for i in self.year_to_count.values()]
        ax.bar(x_list1, vacanciesAvg, width, label="Количество вакансий")
        ax.bar(x_list2, vacancies, width, label="Количество вакансий аналитик")
        ax.set_xticks(x_nums, years, rotation="vertical")
        ax.tick_params(axis="both", labelsize=8)
        ax.legend(fontsize=8)
        ax.grid(True, axis="y")

        ax = fig.add_subplot(223)
        ax.invert_yaxis()
        # ax.invert_xaxis()
        ax.set_title("Уровень зарплат по городам", fontsize=8)
        cyties = [i for i in self.areaSalary.keys()]
        salary = [i for i in self.areaSalary.values()]
        ax.barh(cyties, salary, width, label="уровень з/п", align='center')
        ax.tick_params(axis="both", labelsize=8)
        ax.grid(True, axis="x")

        ax = fig.add_subplot(224)
        ax.set_title("Доля вакансий по городам", fontsize=8)
        cyties = [i for i in self.areaPiece.keys()]
        values = [i for i in self.areaPiece.values()]
        ax.pie(values, labels=cyties)
        ax.tick_params(axis="both", labelsize=8)

        plt.tight_layout()
        plt.savefig("graph.png")
        plt.show()

    def generate_pdf(self):
        '''
        Генерирует пдф отчет
        :return: void
        '''
        env = Environment(loader=FileSystemLoader('.'))
        template = env.get_template('pdf_template.html')
        heads1 = ['Год', 'Средняя зарплата', f'Средняя зарплата - {self.inputValues.professionName}',
                  'Количество вакансий', f'Количество вакансий - {self.inputValues.professionName}']
        heads2 = ['Город', 'Уровень зарплат', 'Город', 'Доля вакансий']
        for key in self.areaPiece.keys():
            self.areaPiece[key] = str(round(self.areaPiece[key] * 100, 2)) + '%'
        pdf_template = template.render({"yearSalary": self.yearSalary,
                                        "yearSalary_needed": self.yearSalary_needed,
                                        "year_to_count": self.year_to_count,
                                        "yearCount": self.yearCount,
                                        "areaSalary": self.areaSalary,
                                        "areaPiece": self.areaPiece,
                                        "heads1": heads1,
                                        "heads2": heads2, })
        config = pdfkit.configuration(wkhtmltopdf=r'C:\Program Files\wkhtmltopdf\bin\wkhtmltopdf.exe')
        pdfkit.from_string(pdf_template, 'report.pdf', configuration=config, options={"enable-local-file-access": None})

class InputConect:
    '''
    Класс для получения входных данных и их валидации
    Attributes:
        fileName (string) имя файла для получения статистики
        professionName (string) название профессии, по которой нужна статистика
    '''
    def __init__(self):
        self.fileName = 'vacancies_dif_currencies.csv'
        self.professionName = 'Аналитик'
        self.areaName = 'Москва'
        self.checkFile()

    def checkFile(self):
        '''
        Валидация входных полей
        :return: void
        '''
        with open(self.fileName, "r", encoding='utf-8-sig', newline='') as csv_file:
            file_iter = iter(csv.reader(csv_file))
            if next(file_iter, "none") == "none":
                print("Пустой файл")
                sys.exit()

            if next(file_iter, "none") == "none":
                print("Нет данных")
                sys.exit()

if __name__ == '__main__':
    # cProfile.run("DataSet()")
    DataSet()