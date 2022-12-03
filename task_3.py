from jinja2 import Environment, FileSystemLoader
import pdfkit
import csv
import math
from openpyxl.worksheet import worksheet
from openpyxl import Workbook
from openpyxl.styles import Border
from openpyxl.styles import Side
from openpyxl.styles import Font
from openpyxl.worksheet.dimensions import DimensionHolder, ColumnDimension
from openpyxl.utils import get_column_letter
import matplotlib.pyplot as plt
from matplotlib.pyplot import figure
import doctest


class PdfReport:
    """
    Класс отчета в виде pdf-файла
    :param years_salary: словарь средних зарплат по годам
    :type years_salary: dict
    :param years_count: словарь количества вакансий по годам
    :type years_count: dict
    :param prof_salary: словарь средних зарплат по годам для введенной профессии
    :type prof_salary: dict
    :param prof_count:словарь количества вакансий по годам для введенной профессии
    :type prof_count: dict
    :param areas_salary:словарь средних зарплат по городам
    :type areas_salary: dict
    :param areas_count: словарь долей от общего количества вакансий по городам
    :type areas_count: dict
    """
    def __init__(self, years_salary: dict, years_count: dict,
                 prof_salary: dict, prof_count: dict,
                 areas_salary: dict, areas_count: dict):
        """
        Инициализирует объект класса пдф отчета
        :param years_salary: словарь средних зарплат по годам
        :type years_salary: dict
        :param years_count: словарь количества вакансий по годам
        :type years_count: dict
        :param prof_salary: словарь средних зарплат по годам для введенной профессии
        :type prof_salary: dict
        :param prof_count:словарь количества вакансий по годам для введенной профессии
        :type prof_count: dict
        :param areas_salary:словарь средних зарплат по городам
        :type areas_salary: dict
        :param areas_count: словарь долей от общего количества вакансий по городам
        :type areas_count: dict
        """
        self.years_salary = years_salary
        self.years_count = years_count
        self.prof_salary = prof_salary
        self.prof_count = prof_count
        self.areas_salary = areas_salary
        self.areas_count = areas_count

    def get_pdf_report(self):
        """
        генерирует отчет в виде пдф-файла на основе html-документа
        :return: pdf-файл
        """
        env = Environment(loader=FileSystemLoader('.'))
        template = env.get_template('pdf_template.html')
        first_header = ["Год", "Средняя зарплата", f"Средняя зарплата {env}", "Количество вакансий", f"Количество вакансий {env}"]
        second_header = ["Город", "Уровень зарплат", "Город", "Доля вакансий"]
        pdf_template = template.render({"year_salary": self.years_salary,
                                       "years_count": self.years_count,
                                       "prof_salary": self.prof_salary,
                                       "prof_count": self.prof_count,
                                       "areas_salary": self.areas_salary,
                                       "areas_count": self.areas_count,
                                        "first_header": first_header,
                                        "second_header": second_header,
                                        "profession": input_set.profession})
        config = pdfkit.configuration(wkhtmltopdf=r"C:\Program Files\wkhtmltopdf\bin\wkhtmltopdf.exe")
        pdfkit.from_string(pdf_template, "report.pdf", configuration=config, options={"enable-local-file-access": ""})


class Salary:
    """
    Класс оклада вакансии
    :param salary_from: Нижняя граница оклада
    :type salary_from: float
    :param salary_to: Верхняя граница оклада
    :type salary_to: float
    :param salary_currency: Валюта оклада
    :type salary_currency: str
    """
    currency_to_rub = {"AZN": 35.68, "BYR": 23.91, "EUR": 59.90, "GEL": 21.74, "KGS": 0.76, "KZT": 0.13, "RUR": 1,
                       "UAH": 1.64, "USD": 60.66, "UZS": 0.0055}

    def __init__(self, salary_from, salary_to, salary_currency):
        """
        Инициализирует объект класса зарплаты
        :param salary_from: Нижняя граница оклада
        :type salary_from: float
        :param salary_to: Верхняя граница оклада
        :type salary_to: float
        :param salary_currency: Валюта оклада
        :type salary_currency: str
        """
        self.salary_from = salary_from
        self.salary_to = salary_to
        self.salary_currency = salary_currency

    def convert_to_rub(self) -> float:
        """
        Конвретирует среднее значение зарплаты в рубли
        :return: зарплата в рублях
        :rtype: float
        """
        value = float(self.currency_to_rub[self.salary_currency])
        return ((float(self.salary_from) + float(self.salary_to)) / 2) * value


class Vacancy:
    """
    Класс вакансии
    :param dict: словарь с данными о вакансии
    :type dict: dict
    :param salary: данные об окладе вакансии
    :type salary: Salary
    """
    def __init__(self, vacancy_dict):
        """
        Инициализирует объект класса вакансии
        :param vacancy_dict: словарь вакансии
        :type vacancy_dict: dict
        """
        self.dict = vacancy_dict
        self.salary = Salary(self.dict['salary_from'], self.dict['salary_to'], self.dict['salary_currency'])


class Input:
    """
    Класс, отвечающий за входные данные
    :param file_name: имя файля
    :type file_name: str
    :param profession: профессия по которой необходимо собирать данные
    :type profession: str
    :param reportType: тип необходимого отчета
    :type reportType: str
    :param fields: поля вакансии
    :type fields: list
    """
    def __init__(self):
        """
        Инициализирует объект класса Input
        """
        self.file_name = input("Введите название файла: ")
        self.profession = input("Введите название профессии: ")
        self.reportType = input("Вакансии или Статистика: ")
        self.fields = []

    def csv_parser(self):
        """
        Парсит csv-файл с вакансиями в массив объектов класса Вакансий
        :return: массив вакансий
        :rtype: list
        """
        reader = self.csv_reader(self.file_name)
        data_all_vacancies = []
        for row in reader:
            new_vacancy = self.get_vacancy(row)
            data_all_vacancies.append(new_vacancy)
        return data_all_vacancies

    def csv_reader(self, file_name) -> []:
        """
        Считывает данные из csv-файла
        :param file_name: имя csv-файла
        :type file_name: str
        :return: массив вакансий в виде массива
        :rtype: list
        """
        with open(file_name, encoding="utf-8-sig") as test:
            unpacker = csv.reader(test)
            data = []
            length = 0
            for row in unpacker:
                if length < len(row):
                    length = len(row)
                if '' not in row and length == len(row):
                    data.append(row)
            self.fields = data[0]
            return data[1:]

    def get_vacancy(self, row: []) -> Vacancy:
        """
        Преобразовывает массив полей вакансии в объект класса Вакансии
        :param row: массив с полями одной вакансии
        :type row: list
        :return: объект класса Вакансии
        :rtype: Vacancy
        """
        vacancy = Vacancy(dict(zip(self.fields, row)))
        return vacancy


class GraphData:
    """
    Класс данных для графиков
    :param data: входные данные о вакансиях
    :type data: list
    :param x_axis: критерий сбора статистики (по годом или по городам)
    :type x_axis: str
    :param profession: профессия по которой необходимо собирать данные
    :type profession: str
    """
    def __init__(self, data, x_axis, profession="not"):
        """
        Инициализирует объект класса GraphData
        :param data: входные данные о вакансиях
        :type data: list
        :param x_axis: критерий сбора статистики (по годом или по городам)
        :type x_axis: str
        :param profession: профессия по которой необходимо собирать данные
        :type profession: str
        """
        self.data = data
        self.profession = profession
        self.salary_data = dict()
        self.count_data = dict()
        self.x_axis = x_axis
        self.calculate_data()

    def calculate_data(self):
        """
        Собирает данные для графиков из входного массива и записывает их в словари
        :return: заполненные словари о зарплатах и количестве вакансий
        """
        for vacancy in self.data:
            self.add_data_from_vacancy(vacancy)
        for x in self.salary_data:
            if self.count_data[x] != 0:
                self.salary_data[x] = math.floor(self.salary_data[x] / self.count_data[x])

    def add_data_from_vacancy(self, vacancy: Vacancy):
        """
        Считывает необходимые данные из одной вакансии в зависимости от критерия и записывает в словарь
        :param vacancy: читаемая вакансия
        :type vacancy: Vacancy
        :return: словари данных для графиков, где учтена входная вакансия
        """
        if self.x_axis == "years":
            abscissa = int(vacancy.dict['published_at'].split('-')[0])
        else:
            abscissa = vacancy.dict['area_name']
        if abscissa not in self.salary_data:
            self.salary_data[abscissa] = 0
        if abscissa not in self.count_data:
            self.count_data[abscissa] = 0
        salary = vacancy.salary.convert_to_rub()
        if self.profession != "not" and self.profession not in vacancy.dict['name']:
            return
        self.update_dicts(abscissa, salary)

    def update_dicts(self, key: str, value: float):
        """
        Обновляет данные в словарях
        :param key: ключ словаря - критерий сбора данных
        :type key: str
        :param value: зарплата вакансии
        :type value: float
        :return: словари с обновленными данными
        """
        try:
            self.salary_data[key] += value
            self.count_data[key] += 1
        except:
            self.salary_data[key] = value
            self.count_data[key] = 1

    def get_graph_data(self):
        """
        Возвращает два словаря с данными для графиков
        Первый словарь - данные по зарплатам
        Второй словарь - данные по количеству вакансий
        :return: Данные для графиков
        :rtype: tuple
        """
        first_printed_dict = self.salary_data
        second_printed_dict = self.count_data
        if self.x_axis == "areas":
            vac_count = sum(list(second_printed_dict.values()))
            first_printed_dict = self.sorted_dict(dict(list(filter(lambda x: self.count_data[x[0]] / vac_count > 0.01, self.salary_data.items()))))
            second_printed_dict = self.sorted_dict(dict(list(filter(lambda x: self.count_data[x[0]] / vac_count > 0.01, self.count_data.items()))))
            for x in second_printed_dict:
                second_printed_dict[x] = float("%.4f" % (second_printed_dict[x] / vac_count))
        return first_printed_dict, second_printed_dict

    @classmethod
    def sorted_dict(cls, non_sorted_dict: dict) -> dict:
        """
        Сортирует словарь по значениям
        :param non_sorted_dict: не отсортированный словарь
        :type non_sorted_dict: dict
        :return: отсортированный словарь
        :rtype: dict
        """
        return dict(list(sorted(non_sorted_dict.items(), key=lambda x: x[1], reverse=True))[:10])


class PngReport:
    """
    Класс для создания отчета в виде диаграмм в png
    :param years_salary: словарь средних зарплат по годам
    :type years_salary: dict
    :param years_count: словарь количества вакансий по годам
    :type years_count: dict
    :param prof_salary: словарь средних зарплат по годам для введенной профессии
    :type prof_salary: dict
    :param prof_count:словарь количества вакансий по годам для введенной профессии
    :type prof_count: dict
    :param areas_salary:словарь средних зарплат по городам
    :type areas_salary: dict
    :param areas_count: словарь долей от общего количества вакансий по городам
    :type areas_count: dict
    """
    def __init__(self, years_salary: dict, years_count: dict,
                 prof_salary: dict, prof_count: dict,
                 areas_salary: dict, areas_count: dict):
        """
        Инициализирует объект класса PngReport
        :param years_salary: словарь средних зарплат по годам
        :type years_salary: dict
        :param years_count: словарь количества вакансий по годам
        :type years_count: dict
        :param prof_salary: словарь средних зарплат по годам для введенной профессии
        :type prof_salary: dict
        :param prof_count:словарь количества вакансий по годам для введенной профессии
        :type prof_count: dict
        :param areas_salary:словарь средних зарплат по городам
        :type areas_salary: dict
        :param areas_count: словарь долей от общего количества вакансий по городам
        :type areas_count: dict
        """
        self.years_salary = years_salary
        self.years_count = years_count
        self.prof_salary = prof_salary
        self.prof_count = prof_count
        self.areas_salary = areas_salary
        self.areas_count = areas_count

    def get_salary_graph(self):
        """
        Генерирует png-файл содержащий статистику по вакансиям в виде диаграмм
        :return: png-файл
        """
        fig = plt.figure()
        plt.rcParams.update({'font.size': 8})
        self.add_bar_subplot(fig, "Уровень зарплат по годам", 221, self.years_salary, "средняя з/п", "y",
                                     subplot_type="", prof_dict=self.prof_salary,
                                     x2_label=f'з/п {input_set.profession}')
        self.add_bar_subplot(fig, "Количество вакансий по годам", 222, self.years_count, "Количество вакансий", "y",
                                     subplot_type="", prof_dict=self.prof_count,
                                     x2_label=f'Количество вакансий {input_set.profession}')
        self.add_bar_subplot(fig, "Уровень зарплат по городам", 223, self.areas_salary, "уровень з/п", "x",
                                     subplot_type="horizontal")
        self.add_pie_sublot(fig, "Доля вакансий по городам", 224, self.areas_count)

        plt.tight_layout()
        plt.savefig("graph.png", dpi=300)

    @classmethod
    def add_bar_subplot(cls, fig: figure, title: str, width: int, full_dict: dict, x1_label: str,  axis, subplot_type="", prof_dict={}, x2_label=""):
        """
        Добавляет столбчатую диаграмму в отчет
        :param fig: область в которую добавляются диаграмма
        :type fig: figure
        :param title: название диаграммы
        :type title: str
        :param width: ширина диаграммы
        :type width: int
        :param full_dict: словарь данных по всем вакансиям для диаграммы
        :type full_dict: dict
        :param x1_label: заголовок для полного словаря данных
        :type x1_label: str
        :param axis: ось на которой располагаются столбцы диаграммы
        :type axis: str
        :param subplot_type: тип столбчатый диаграммы (горизонтальный или вертикальный)
        :type subplot_type: str
        :param prof_dict: словарь данных для диаграммы по вакансиям для введенной профессии
        :type prof_dict: dict
        :param x2_label: заголовок для диаграммы по данным о введенной профессии
        :type x2_label: str
        :return: на область добавлена диагамма
        """
        ax = fig.add_subplot(width)
        ax.set_title(title, fontsize=8)
        ax.tick_params(axis="both", labelsize=8)
        ax.grid(True, axis=axis)
        if subplot_type == "horizontal":
            ax.barh(list(full_dict.keys()), list(full_dict.values()), label=x1_label, align="center")
            ax.invert_yaxis()
        else:
            x_axis = range(len(full_dict.keys()))
            x1 = list(map(lambda x: float(x) - 0.2, x_axis))
            x2 = list(map(lambda x: float(x) + 0.2, x_axis))
            ax.bar(x1, list(full_dict.values()), width=0.4, label=x1_label)
            ax.bar(x2, list(prof_dict.values()), width=0.4, label=x2_label)
            ax.set_xticks(x_axis, list(full_dict.keys()), rotation="vertical")
        ax.legend(fontsize=8)

    @classmethod
    def add_pie_sublot(cls, fig: figure, title: str, width: int, data: dict):
        """
        Добавляет круговую диаграмму в отчет
        :param fig: Область в которую добавляется диаграмма
        :type fig: figure
        :param title: Заголовок диаграммы
        :type title: str
        :param width: Ширина диаграммы
        :type width: int
        :param data: Словарь данных для диаграммы
        :type data: dict
        :return: на область добавлена круговая диаграмма
        """
        ax = fig.add_subplot(width)
        ax.set_title(title, fontsize=8)
        ax.tick_params(axis="both", labelsize=8)
        keys = list(data.keys())
        values = list(data.values())
        ax.pie(values, labels=keys)


class ExcelReport:
    """
    Класса для создания отчета в виде Excel таблиц
    :param border: границы ячеек в таблице
    :type border: Border
    :param font: шрифт текста в таблице
    :type font: Font
    """
    def __init__(self, side, font):
        """
        Инициализирует объект класса ExcelReport
        :param side: тип границ для ячеек таблицы
        :type side: Side
        :param font: шрифт текста в отчете
        :type font: Font
        """
        self.border = Border(left=side, top=side, right=side, bottom=side)
        self.font = font

    def generate_excel(self, dicts: list):
        """
        Генерирует excel-файл со статистикой собранной из входных данных
        :param dicts: вхожные данные
        :type dicts: list
        :return: сгенерированный excel-файл со статистикой в виде таблиц
        """
        wb = Workbook()
        years_sheet = self.generate_years_sheet(wb, dicts[0], dicts[1], dicts[2], dicts[3])
        areas_sheet = self.generate_areas_sheet(wb, dicts[4], dicts[5])
        ws = wb.worksheets[0]
        wb.remove(ws)
        wb.save("report.xlsx")

    def generate_years_sheet(self, wb: Workbook, salaries: dict, prof_salaries: dict, counts: dict, prof_counts: dict) -> worksheet.Worksheet:
        """
        Генирирует лист в excel-книге со статистикой по годам
        :param wb: excel-книга
        :type wb: Workbook
        :param salaries: словарь средних зарплат по годам
        :type salaries: dict
        :param prof_salaries: словарь средних зарплат по годам для введенной профессии
        :type prof_salaries: dict
        :param counts: словарь количества вакансий по годам
        :type counts: dict
        :param prof_counts: словарь количества вакансий по годам для введенной профессии
        :type prof_counts: dict
        :return: excel-лист со статистикой по годам
        """
        years_sheet = wb.create_sheet("Статистика по годам")
        years_sheet['A1'] = 'Год'
        years_sheet['B1'] = 'Средняя зарплата'
        years_sheet['C1'] = f'Средняя зарплата - {input_set.profession}'
        years_sheet['D1'] = 'Количество вакансий'
        years_sheet['E1'] = f'Количество вакансий - {input_set.profession}'
        column_cells = [years_sheet["A1"], years_sheet['B1'], years_sheet['C1'], years_sheet['D1'], years_sheet['E1']]
        for cell in column_cells:
            cell.border = self.border
            cell.font = self.font
        for x in range(len(salaries)):
            years_sheet[f'A{x + 2}'] = list(salaries.keys())[x]
            years_sheet[f'B{x + 2}'] = list(salaries.values())[x]
            years_sheet[f'C{x + 2}'] = list(prof_salaries.values())[x]
            years_sheet[f'D{x + 2}'] = list(counts.values())[x]
            years_sheet[f'E{x + 2}'] = list(prof_counts.values())[x]
            years_sheet[f'A{x + 2}'].border = self.border
            years_sheet[f'B{x + 2}'].border = self.border
            years_sheet[f'C{x + 2}'].border = self.border
            years_sheet[f'D{x + 2}'].border = self.border
            years_sheet[f'E{x + 2}'].border = self.border
        dim_holder = DimensionHolder(worksheet=years_sheet)

        for col in range(years_sheet.min_column, years_sheet.max_column + 1):
            dim_holder[get_column_letter(col)] = ColumnDimension(years_sheet, min=col, max=col, width=20)

        years_sheet.column_dimensions = dim_holder

        return years_sheet

    def generate_areas_sheet(self, wb: Workbook, salaries: dict, counts: dict) -> worksheet.Worksheet:
        """
        Генирирует лист в excel-книге со статистикой по городам
        :param wb: excel-книга
        :type wb: Workbook
        :param salaries:словарь средних зарплат по городам
        :type salaries: dict
        :param counts: словарь долей от общего количества вакансий по городам
        :type counts: dict
        :return: excel-лист со статистикой по городам
        """
        areas_sheet = wb.create_sheet("Статистика по городам")
        column_cells = [areas_sheet['A1'], areas_sheet['B1'], areas_sheet['D1'], areas_sheet['E1']]
        for cell in column_cells:
            cell.border = self.border
            cell.font = self.font
        areas_sheet['A1'] = "Город"
        areas_sheet['B1'] = "Уровень зарплат"
        areas_sheet['D1'] = "Город"
        areas_sheet['E1'] = "Доля вакансий"
        for x in range(10):
            areas_sheet[f'A{x + 2}'] = list(salaries.keys())[x]
            areas_sheet[f'B{x + 2}'] = list(salaries.values())[x]
            areas_sheet[f'D{x + 2}'] = list(counts.keys())[x]
            areas_sheet[f'E{x + 2}'] = list(counts.values())[x]
            areas_sheet[f'A{x + 2}'].border = self.border
            areas_sheet[f'B{x + 2}'].border = self.border
            areas_sheet[f'D{x + 2}'].border = self.border
            areas_sheet[f'E{x + 2}'].border = self.border
        dim_holder = DimensionHolder(worksheet=areas_sheet)

        for col in range(areas_sheet.min_column, areas_sheet.max_column + 1):
            dim_holder[get_column_letter(col)] = ColumnDimension(areas_sheet, min=col, max=col, width=20)

        areas_sheet.column_dimensions = dim_holder
        return areas_sheet


"Вводим данные"
input_set = Input()


def main():
    """
    Главный входной метод программы
    Генерирует файл с отчетом, тип отчета зависит от выбора пользователя
    :return: Сгенирированный файл отчета
    """
    data = input_set.csv_parser()
    years = GraphData(data, "years")
    prof_years = GraphData(data, "years", input_set.profession)
    areas = GraphData(data, "areas")
    if input_set.reportType == "Вакансии":
        reportExcel = ExcelReport(Side(style="thin", color="000000"), Font(bold=True))
        reportExcel.generate_excel([years.get_graph_data()[0], years.get_graph_data()[1],
                                    prof_years.get_graph_data()[0], prof_years.get_graph_data()[1],
                                    areas.get_graph_data()[0], areas.get_graph_data()[1]])
    elif input_set.reportType == "Статистика":
        pdf_report = PdfReport(years.get_graph_data()[0], years.get_graph_data()[1],
                                    prof_years.get_graph_data()[0], prof_years.get_graph_data()[1],
                                    areas.get_graph_data()[0], areas.get_graph_data()[1])
        pdf_report.get_pdf_report()


"Запускаем программу"
main()




